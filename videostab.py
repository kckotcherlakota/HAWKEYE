#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Sep 28 16:15:45 2019

@author: krishna
"""
import face_recognition
import cv2
import numpy as np
#
from vidstab import VidStab
from openpyxl import Workbook
import datetime


book=Workbook()
sheet=book.active

now= datetime.datetime.now()
today=now.day
month=now.month


video_capture = cv2.VideoCapture(0)
pavan_img = face_recognition.load_image_file("vineet.jpeg")
pavan = face_recognition.face_encodings(pavan_img)[0]
akhil_img = face_recognition.load_image_file("akhil.jpeg")
akhil = face_recognition.face_encodings(akhil_img)[0]
#sahithya_img = face_recognition.load_image_file("sahithya.jpeg")
#sahithya = face_recognition.face_encodings(sahithya_img)[0]
known_face_encodings = [pavan, akhil]
known_face_names = ["1","2"]
known_face_star=["vinnet","akhil"]
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True
stabilizer = VidStab(kp_method='FAST', threshold=42, nonmaxSuppression=True)

while True:
    ret, frame = video_capture.read()
   
    stabilized_frame = stabilizer.stabilize_frame(input_frame=frame, border_size=50) 
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = small_frame[:, :, ::-1]
    if process_this_frame:

        face_locations = face_recognition.face_locations(rgb_small_frame,model="cnn")
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:

            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_face_names[best_match_index]
                x=1
                if int(name) in range(1,10):
                    sheet.cell(row=int(name)+3, column=1).value = "not found"
                    sheet.cell(row=1, column=1).value="name of the criminal"
                    sheet.cell(row=1, column=2).value = "location of the criminal on "+str(today)+"-"+str(month)+"-19"
                while x<=10:
                    if sheet.cell(row=x+2, column=1).value !="not found":
                        sheet.cell(row=x+2, column=1).value =known_face_star[int(name)]
                        sheet.cell(row=x+2, column=2).value="found at SRM University AP"
                        x=x+1
                    else:
                        break
#                   
#                
                    
                
#                sheet.cell(row=2, column=1).value="absent"
#                
                

            face_names.append(name)

    process_this_frame = not process_this_frame


    for (top, right, bottom, left), name in zip(face_locations, face_names):

        top *= 4
        right *= 4
        bottom *= 4
        left *= 4


        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)


        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)


    cv2.imshow('Video', frame)
    book.save('Don.xlsx')
    


    if cv2.waitKey(1) & 0xFF == ord('c'):
        break
video_capture.release()
cv2.destroyAllWindows()